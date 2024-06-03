
from flask import Blueprint, render_template, request, flash, jsonify, redirect, url_for,send_from_directory
from flask_login import login_required, current_user
from .models import Event,Payment, MapMarker, User, Message,Report, Survey, Answer, Question, CATEGORIES,segregate_waste,Voucher,stores
from datetime import datetime, date
from . import db
from . import mail
from flask_mail import Message as MailMessage
import json,calendar,stripe,random,string,requests,uuid
import os
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook

views = Blueprint('views', __name__)

polish_month_names = {
    1: "Styczeń", 2: "Luty", 3: "Marzec", 4: "Kwiecień", 5: "Maj", 6: "Czerwiec",
    7: "Lipiec", 8: "Sierpień", 9: "Wrzesień", 10: "Październik", 11: "Listopad", 12: "Grudzień"
}

def send_confirmation_email(user):
    msg = MailMessage('Stworzono konto w systemie E-Gmina', recipients=[user.email])
    msg.body = f'''Dziękujemy za założenie konta na E-Gmina
    '''
    mail.send(msg)

def get_month_events(year, month):
    start_date = date(year, month, 1)
    end_date = date(year, month, calendar.monthrange(year, month)[1])
    month_events = Event.query.filter(Event.date.between(start_date, end_date)).all()
    return month_events


def generate_calendar(year, month, month_events):
    cal = calendar.monthcalendar(year, month)
    weeks = []
    for week in cal:
        current_week = []
        for day in week:
            events_for_day = [event for event in month_events if event.date.day == day]
            current_week.append((day, events_for_day))
        weeks.append(current_week)
    return weeks

@views.route('/', methods=['GET', 'POST'])
def home():
    today = date.today()
    today_year = today.year
    today_month = today.month

    # Domyślnie ustaw na bieżący miesiąc i rok
    year = today_year
    month = today_month

    # Sprawdź parametry GET
    if 'current_month' in request.args and 'current_year' in request.args:
        try:
            month = int(request.args.get('current_month'))
            year = int(request.args.get('current_year'))
        except ValueError:
            # Błędne wartości, użyj domyślnych
            month = today_month
            year = today_year

    # Oblicz poprzedni i następny miesiąc i rok
    prev_month = month - 1
    next_month = month + 1
    prev_year = year
    next_year = year

    if prev_month == 0:
        prev_month = 12
        prev_year -= 1
    if next_month == 13:
        next_month = 1
        next_year += 1

    month_events = get_month_events(year, month)
    accepted_month_events = [event for event in month_events if event.status == 'accepted']
    weeks = generate_calendar(year, month, accepted_month_events)
    # Pobierz tylko przyszłe zaakceptowane wydarzenia i sortuj je chronologicznie
    accepted_events = Event.query.filter(Event.date >= datetime.now(), Event.status == 'accepted').order_by(Event.date.asc()).all()
    markers = MapMarker.query.all()  # Pobieramy wszystkie znaczniki z bazy danych

    # Pobierz nazwę aktualnego miesiąca po polsku
    current_month_name = polish_month_names[month]

    return render_template("home.html", calendar=weeks, current_month=current_month_name, current_year=year,
                           today_month=today_month, today_year=today_year,
                           prev_month=prev_month, prev_year=prev_year,
                           next_month=next_month, next_year=next_year,
                           user=current_user, accepted_events=accepted_events, markers=markers)



@views.route('/delete-event', methods=['POST'])
def delete_event():  
    event = json.loads(request.data)
    eventId = event['eventId']
    event = Event.query.get(eventId)
    if event:
        if event.user_id == current_user.id or current_user.role == "admin":
            db.session.delete(event)
            db.session.commit()

    return jsonify({})


@views.route('/calendar', methods=['GET'])
def calendar_view():
    today = datetime.date.today()
    current_month = today.strftime("%B")
    year = today.year
    month = today.month
    month_events = get_month_events(year, month)
    accepted_month_events = [event for event in month_events if event.status == 'accepted']
    weeks = generate_calendar(year, month, accepted_month_events)
    return render_template("calendar.html", user=current_user, calendar=weeks, current_month=current_month)


@views.route('/maps', methods=['GET', 'POST'])
def maps():
    if request.method == 'POST':
        address = request.form['address']
        description = request.form['description']
        api_key = 'AIzaSyDgRv7f0CZS1zchzAV9WsXTMRrCmIHxY_M'
        geocoding_url = f'https://maps.googleapis.com/maps/api/geocode/json?address={address}&key={api_key}'

        response = requests.get(geocoding_url)
        data = response.json()

        if data['status'] == 'OK':
            lat = data['results'][0]['geometry']['location']['lat']
            lng = data['results'][0]['geometry']['location']['lng']

            new_marker = MapMarker(lat=lat, lng=lng, address=address, description=description, user_id=current_user.id)
            db.session.add(new_marker)
            db.session.commit()

            flash('Znacznik dodany pomyślnie', category='success')
            return redirect(url_for('views.maps'))
        else:
            flash('Nie znaleziono koordynatów dla podanego adresu', category='error')
            return redirect(url_for('views.maps'))

    markers = MapMarker.query.all()  # Pobieramy wszystkie znaczniki z bazy danych
    return render_template("maps.html", user=current_user, markers=markers)

@views.route('/delete-marker/<int:marker_id>', methods=['POST'])
def delete_marker(marker_id):
    marker = MapMarker.query.get(marker_id)
    if marker:
        db.session.delete(marker)
        db.session.commit()
        flash('Znacznik usunięty pomyślnie', category='success')
    else:
        flash('Znacznik nie został znaleziony', category='error')
    return redirect(url_for('views.maps'))


@views.route('/edit-marker', methods=['POST'])
def edit_marker():
    marker_id = request.form['editMarkerId']
    new_address = request.form['editAddress']
    new_description = request.form['editDescription']
    api_key = 'AIzaSyDgRv7f0CZS1zchzAV9WsXTMRrCmIHxY_M'
    geocoding_url = f'https://maps.googleapis.com/maps/api/geocode/json?address={new_address}&key={api_key}'

    response = requests.get(geocoding_url)
    data = response.json()

    marker = MapMarker.query.get(marker_id)
    if data['status'] == 'OK':
        lat = data['results'][0]['geometry']['location']['lat']
        lng = data['results'][0]['geometry']['location']['lng']

        new_marker = MapMarker(lat=lat, lng=lng, address=new_address, description=new_description, user_id=current_user.id)
        db.session.delete(marker)
        db.session.add(new_marker)
        db.session.commit()
        flash('Znacznik zaktualizowany pomyślnie', category='success')
    else:
        flash('Nie znaleziono znacznika do edycji', category='error')

    return redirect(url_for('views.maps'))


@views.route('/events', methods=['GET', 'POST'])
@login_required
def event():
    if current_user.role=="admin":
        user_events = Event.query.order_by(Event.date.asc()).all()
    else:
        user_events = Event.query.filter(Event.user_id == current_user.id, Event.date >= datetime.now()).order_by(Event.date.asc()).all()
    if request.method == 'POST':
        data = request.form.get('data')
        date = request.form.get('date')
        place = request.form.get('place')
        name = request.form.get('name')

        if data and date and place and name:
            try:
                date = datetime.strptime(date, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Zły format daty', category='error')
                return redirect(url_for('views.event'))

            # Sprawdzamy poprawność adresu za pomocą API Google Maps
            api_key = 'AIzaSyDgRv7f0CZS1zchzAV9WsXTMRrCmIHxY_M'
            geocoding_url = f'https://maps.googleapis.com/maps/api/geocode/json?address={place}&key={api_key}'
            response = requests.get(geocoding_url)
            data = response.json()

            if data['status'] == 'OK':
                if current_user.role =="admin":
                    status_state = "accepted"
                else:
                    status_state = "pending"
                new_event = Event(
                    data=request.form.get('data'),  
                    date=date,
                    place=place,
                    name=name,
                    user_id=current_user.id,
                    status=status_state
                )
                db.session.add(new_event)
                db.session.commit()
                flash('Wydarzenie zawnioskowano!', category='success')
                return redirect(url_for('views.event'))
            else:
                flash('Nie znaleziono koordynatów dla podanego adresu', category='error')
                return redirect(url_for('views.event'))

    return render_template('events.html', user=current_user, user_events=user_events)


@views.route('/invite-friends', methods=['GET', 'POST'])
@login_required
def search_user():
    if request.method == 'POST':
        username = request.form.get('username')
        user = User.query.filter_by(username=username).first()
        if user:
            if user == current_user:
                flash('Nie możesz dodać samego siebie jako znajomego!', category='error')
            else:
                if user in current_user.friends:
                    flash(f'{user.username} jest już twoim znajomym!', category='error')
                elif user in current_user.sent:
                    flash(f'Zaproszenie wysłano do {user.username}!', category='error')
                elif user in current_user.invitations:
                    flash(f'Zaproszenie otrzymane od {user.username}!', category='error')
                elif user in current_user.blocked:
                    flash(f'{user.username} jest zablokowany!', category='error')
                elif user in current_user.y_blocked:
                    flash(f'Zostałeś zablokowany przez {user.username}!', category='error')
                elif user.role == 'admin':
                    flash(f'Nie można wysłać wiadomości do admina!', category='error')
                elif user.role == 'organisation':
                    flash(f'Nie można wysłać wiadomości do organizacji!', category='error')
                else:
                    current_user.sent.append(user)
                    user.invitations.append(current_user)
                    db.session.commit()
                    flash(f'Zaproszenie pomyślnie wysłano do {user.username}!', category='success')
        else:
            flash('Użytkownik nie został znaleziony!', category='error')
        return redirect(url_for('views.search_user'))
    return render_template('invite-friends.html', user=current_user)


@views.route('/send-message/<int:user_id>', methods=['POST'])
@login_required
def send_message(user_id):
    if request.method == 'POST':
        content = request.form.get('content')
        receiver = User.query.get(user_id)
        if receiver:
            new_message = Message(sender_id=current_user.id, receiver_id=user_id, content=content)
            db.session.add(new_message)
            db.session.commit()
            flash('Wiadomość została wysłana!', category='success')
            return redirect(url_for('views.user_chat', user_id=user_id))
        else:
            flash('Nie znaleziono użytkownika.', category='error')
    return redirect(url_for('views.home'))


@views.route('/remove-friend/<int:friend_id>', methods=['POST'])
@login_required
def remove_friend(friend_id):
    friend = User.query.get(friend_id)
    if friend:
        current_user.friends.remove(friend)
        friend.friends.remove(current_user)
        db.session.commit()
        flash(f'{friend.username} został usunięty z listy znajomych.', category='success')
    else:
        flash('Nie znaleziono użytkownika.', category='error')
    return redirect(url_for('views.search_user'))


@views.route('/block-user/<int:user_id>', methods=['POST'])
@login_required
def block_user(user_id):
    user = User.query.get(user_id)
    if user:
        if user in current_user.friends:
            current_user.blocked.append(user)
            user.y_blocked.append(current_user)
            current_user.friends.remove(user)
            user.friends.remove(current_user)
            db.session.commit()
            flash(f'{user.username} został zablokowany.', category='success')
        else:
            current_user.blocked.append(user)
            user.y_blocked.append(current_user)
            current_user.invitations.remove(user)
            user.sent.remove(current_user)
            db.session.commit()
            flash(f'{user.username} został zablokowany.', category='success')
    else:
        flash('Nie znaleziono użytkownika.', category='error')
    return redirect(url_for('views.search_user'))


@views.route('/unblock-user/<int:user_id>', methods=['POST'])
@login_required
def unblock_user(user_id):
    user = User.query.get(user_id)
    if user:
        if user in current_user.blocked:
            current_user.blocked.remove(user)
            user.y_blocked.remove(current_user)
            db.session.commit()
            flash(f'{user.username} został odblokowany.', category='success')
        else:
            flash('Użytkownik nie jest zablokowany.', category='error')
    else:
        flash('Nie znaleziono użytkownika.', category='error')
    return redirect(url_for('views.search_user'))


@views.route('/chat/<int:user_id>', methods=['GET'])
@login_required
def user_chat(user_id):
    receiver = User.query.get(user_id)
    if not receiver:
        flash('Nie znaleziono użytkownika.', category='error')
        return redirect(url_for('views.home'))
    #username = request.form.get('username')
    #user = User.query.filter_by(username=username).first()
    elif receiver in current_user.y_blocked:
        flash(f'Zostałeś zablokowany przez {receiver.username}!', category='error')
        return redirect(url_for('views.search_user'))
    elif receiver not in current_user.friends:
        flash(f'{receiver.username} nie jest już twoim znajomym!', category='error')
        return redirect(url_for('views.search_user'))
    messages_sent_by_user = Message.query.filter_by(sender_id=current_user.id, receiver_id=user_id).all()
    messages_received_by_user = Message.query.filter_by(sender_id=user_id, receiver_id=current_user.id).all()
    messages = messages_sent_by_user + messages_received_by_user
    messages.sort(key=lambda x: x.timestamp)

    return render_template('user_chat.html', user=current_user, receiver=receiver, messages=messages)


@views.route('/accept-invite', methods=['POST'])
@login_required
def accept_invite():
    data = json.loads(request.data)
    inviter_email = data['inviter_email']
    inviter = User.query.filter_by(username=inviter_email).first()
    if inviter in current_user.invitations:
        current_user.invitations.remove(inviter)
        current_user.friends.append(inviter)
        inviter.friends.append(current_user)
        inviter.sent.remove(current_user)
        db.session.commit()
    return jsonify({})


@views.route('/reject-invite', methods=['POST'])
@login_required
def reject_invite():
    data = json.loads(request.data)
    inviter_email = data['inviter_email']
    inviter = User.query.filter_by(username=inviter_email).first()
    if inviter in current_user.invitations:
        current_user.invitations.remove(inviter)
        inviter.sent.remove(current_user)
        db.session.commit()
    return jsonify({})

@views.route('/admin-events', methods=['GET', 'POST'])
@login_required
def admin_events():
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej strony.', category='danger')
        return redirect(url_for('views.home'))
    
    pending_events = Event.query.filter_by(status='pending').all()
    
    if request.method == 'POST':
        event_id = request.form.get('event_id')
        action = request.form.get('action')  
        
        event = Event.query.get(event_id)
        if not event:
            flash('Nie znaleziono wydarzenia.', category='danger')
            return redirect(url_for('views.admin_events'))
        
        if action == 'accept':
            event.status = 'accepted'
            db.session.add(event)
            db.session.commit()
            flash('Wydarzenie zostało zaakceptowane.', category='success')
        elif action == 'reject':
            event.status = 'rejected'
            db.session.add(event)
            db.session.commit()
            flash('Wydarzenie zostało odrzucone.', category='warning')
        
        return redirect(url_for('views.admin_events'))

    return render_template('admin_events.html', user=current_user, pending_events=pending_events)

@views.route('/update_event_status/<int:event_id>/<string:action>', methods=['POST'])
@login_required
def update_event_status(event_id, action):
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej akcji.', category='danger')
        return redirect(url_for('views.admin_events'))

    event = Event.query.get(event_id)
    if not event:
        flash('Nie znaleziono wydarzenia.', category='danger')
        return redirect(url_for('views.admin_events'))

    if action == 'accept':
        event.status = 'accepted'
        user = User.query.filter_by(id=event.user_id).first()
        add_loyalty_points(user, 10)
        db.session.commit()
        flash('Wydarzenie zostało zaakceptowane.', category='success')
    elif action == 'reject':
        event.status = 'rejected'
        db.session.commit()
        flash('Wydarzenie zostało odrzucone.', category='warning')

    return redirect(url_for('views.admin_events'))


@views.route('/admin-users', methods=['GET', 'POST'])
@login_required
def admin_users():
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej strony.', category='danger')
        return redirect(url_for('views.home'))

    pending_users = User.query.filter_by(status='pending').all()
    accepted_users = User.query.filter_by(status='accepted').all()

    # Wykluczenie konta administratora
    accepted_users = [user for user in accepted_users if user.role != 'admin']

    if request.method == 'POST':
        user_id = request.form.get('user_id')
        action = request.form.get('action')

        user = User.query.get(user_id)
        if user:
            if action == 'accept':
                user.status = 'accepted'
                db.session.add(user)
                db.session.commit()
                flash('Stworzenie konta zostało zatwierdzone.', category='success')
            elif action == 'reject':
                user.status = 'rejected'
                db.session.commit()
                flash('Stworzenie konta zostało odrzucone.', category='warning')

        return redirect(url_for('views.admin_users'))

    return render_template('admin_users.html', user=current_user, pending_users=pending_users,
                           accepted_users=accepted_users)


@views.route('/delete-user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej operacji.', category='danger')
        return redirect(url_for('views.home'))

    user = User.query.get(user_id)
    if user:
        if user.id == current_user.id:
            flash('Nie możesz usunąć samego siebie.', category='danger')
        else:
            db.session.delete(user)
            db.session.commit()
            flash('Użytkownik został usunięty.', category='success')
    else:
        flash('Użytkownik nie istnieje.', category='danger')

    return redirect(url_for('views.admin_users'))

@views.route("/report", methods=['GET', 'POST'])
@login_required
def report():
    user_reports = list(reversed(current_user.Reports))
    if request.method == 'POST':
        data = request.form.get('data')
        place = request.form.get('place')
        date = datetime.now()

        if data and place:
            # Sprawdzamy poprawność adresu za pomocą API Google Maps
            api_key = 'AIzaSyDgRv7f0CZS1zchzAV9WsXTMRrCmIHxY_M'
            geocoding_url = f'https://maps.googleapis.com/maps/api/geocode/json?address={place}&key={api_key}'
            response = requests.get(geocoding_url)
            response_data = response.json()

            if response_data['status'] == 'OK' and response_data['results']:
                # Adres jest poprawny, zapisujemy zgłoszenie
                new_report = Report(
                    data=data,
                    date=date,
                    place=place,
                    user_id=current_user.id,
                    status='pending'
                )
                db.session.add(new_report)
                db.session.commit()
                flash('Raport zawnioskowano!', category='success')
                return redirect(url_for('views.report'))
            else:
                flash('Nie znaleziono koordynatów dla podanego adresu', category='error')
        else:
            flash('Wszystkie pola muszą być wypełnione', category='error')

    all_hidden = all(report.hidden for report in user_reports)
    return render_template('report.html', user=current_user, user_reports=user_reports, all_hidden=all_hidden)


@views.route("/report/delete/<int:report_id>", methods=['POST'])
@login_required
def delete_report(report_id):
    report_to_delete = Report.query.get_or_404(report_id)
    if report_to_delete.user_id == current_user.id:
        db.session.delete(report_to_delete)
        db.session.commit()
        flash('Zgłoszenie zostało usunięte.', category='success')
    else:
        flash('Nie masz uprawnień do usunięcia tego zgłoszenia.', category='error')
    return redirect(url_for('views.report'))

@views.route("/report/hide/<int:report_id>", methods=['POST'])
@login_required
def hide_report(report_id):
    report_to_hide = Report.query.get_or_404(report_id)
    if report_to_hide.user_id == current_user.id:
        report_to_hide.hidden = True
        db.session.commit()
        flash('Zgłoszenie zostało ukryte.', category='success')
    else:
        flash('Nie masz uprawnień do ukrycia tego zgłoszenia.', category='error')
    return redirect(url_for('views.report'))


@views.route('/update_user_status/<int:user_id>/<string:action>', methods=['POST'])
@login_required
def update_user_status(user_id, action):
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej akcji.', category='danger')
        return redirect(url_for('views.admin_users'))

    user = User.query.get(user_id)
    if not user:
        flash('Nie znaleziono użytkownika.', category='danger')
        return redirect(url_for('views.admin_users'))

    if action == 'accept':
        user.status = 'accepted'
        db.session.commit()
        flash('Użytkownik został zaakceptowany.', category='success')
    elif action == 'reject':
        user.status = 'rejected'
        db.session.commit()
        flash('Użytkownik został odrzucony.', category='warning')

    return redirect(url_for('views.admin_users'))

@views.route('/update_report_status/<int:report_id>/<string:action>', methods=['POST'])
@login_required
def update_report_status(report_id, action):
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej akcji.', category='danger')
        return redirect(url_for('views.admin_reports'))

    report = Report.query.get(report_id)
    if not report:
        flash('Nie znaleziono zgłoszenia.', category='danger')
        return redirect(url_for('views.admin_reports'))

    if action == 'accept':
        report.status = 'accepted'
        user = User.query.filter_by(id=report.user_id).first()
        add_loyalty_points(user, 10)
        db.session.commit()
        flash('Zgłoszenie zostało zaakceptowane.', category='success')
    elif action == 'reject':
        report.status = 'rejected'
        db.session.commit()
        flash('Zgłoszenie zostało odmówione.', category='warning')

    return redirect(url_for('views.admin_reports'))

@views.route('/admin-reports', methods=['GET', 'POST'])
@login_required
def admin_reports():
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Nie masz uprawnień administratora do tej strony.', category='danger')
        return redirect(url_for('views.home'))
    
    pending_reports = Report.query.filter_by(status='pending').all()
    
    if request.method == 'POST':
        report_id = request.form.get('report_id')
        action = request.form.get('action')  
        
        report = Report.query.get(report_id)
        if not report:
            flash('Nie znaleziono zgłoszenia.', category='danger')
            return redirect(url_for('views.admin_reports'))
        
        if action == 'accept':
            report.status = 'accepted'
            db.session.add(report)
            db.session.commit()
            flash('Zgłoszenie zostało zaakceptowane.', category='success')
        elif action == 'reject':
            report.status = 'rejected'
            db.session.add(report)
            db.session.commit()
            flash('Zgłoszenie zostało odrzucone.', category='warning')
        
        return redirect(url_for('views.admin_reports'))

    return render_template('admin_reports.html', user=current_user, pending_reports=pending_reports)


@views.route('/surveys', methods=['GET'])
@login_required
def surveys():
    surveys = Survey.query.filter_by(status='active').all()
    return render_template('surveys.html', surveys=surveys,user=current_user)

@views.route('/survey/<int:survey_id>', methods=['GET', 'POST'])
@login_required
def fill_survey(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    if request.method == 'POST':
        for question in survey.questions:
            answer_text = request.form.get(f'question_{question.id}')
            answer = Answer(text=answer_text, question_id=question.id, user_id=current_user.id)
            db.session.add(answer)
        db.session.commit()
        survey.status = "completed"
        add_loyalty_points(current_user, 5)
        db.session.add(survey)
        db.session.commit()
        flash('Dziękujemy za wypełnienie ankiety!', 'success')
        return redirect(url_for('views.surveys'))
    return render_template('survey.html', survey=survey, user=current_user)

@views.route('/create-survey', methods=['GET', 'POST'])
@login_required
def create_survey():
    if request.method == 'POST':
        title = request.form.get('title')
        survey = Survey(title=title)
        questions = request.form.getlist('questions[]')
        db.session.add(survey)
        db.session.commit()
        for question_text in questions:
            if question_text.strip():  
                question = Question(text=question_text, survey_id=survey.id)
                db.session.add(question)

        db.session.commit()  
        flash('Ankieta została utworzona!', 'success')
        return redirect(url_for('views.create_survey'))
    return render_template('create_survey.html',user=current_user)

@views.route('/segregate', methods=['GET', 'POST'])
@login_required
def segregate():
    result = None
    if request.method == 'POST':
        description = request.form.get('waste_description')
        result = segregate_waste(description)
    
    return render_template('segregate.html', result=result, user=current_user,categories=CATEGORIES)


@views.route('/user_info')
@login_required
def user_info():
    return render_template('user_info.html', user=current_user)

def add_loyalty_points(user, points):
    user.loyalty_points += points
    db.session.add(user)
    db.session.commit()
    
@views.route('/exchange', methods=['GET', 'POST'])
@login_required
def exchange_points():
    available_stores = stores

    if request.method == 'POST':
        if current_user.role == "admin":
            if 'new_store_name' in request.form:
                new_store_name = request.form.get('new_store_name')
                new_store_cost = int(request.form.get('new_store_cost'))
                stores[f'store{len(stores) + 1}'] = {'name': new_store_name, 'cost': new_store_cost}
                flash('Nowy sklep został dodany!', 'success')
            elif 'delete_store' in request.form:
                store_key = request.form.get('delete_store')
                if store_key in stores:
                    del stores[store_key]
                    flash('Sklep został usunięty!', 'success')
                else:
                    flash('Sklep nie istnieje.', 'danger')
        else:
            store_key = request.form.get('store')
            store = available_stores.get(store_key)
            if store and current_user.loyalty_points >= store['cost']:
                voucher_code = str(uuid.uuid4()).replace('-', '').upper()[:12]
                new_voucher = Voucher(user_id=current_user.id, store_name=store['name'], code=voucher_code)
                current_user.loyalty_points -= store['cost']
                db.session.add(new_voucher)
                db.session.commit()
                flash(f'Wymieniono punkty na bon do {store["name"]}! Kod: {voucher_code}', 'success')
            else:
                flash('Nie masz wystarczającej liczby punktów lub nieprawidłowy sklep.', 'danger')
        return redirect(url_for('views.exchange_points'))

    return render_template('exchange_points.html', user=current_user, stores=stores)

@views.route('/my-vouchers', methods=['GET'])
@login_required
def my_vouchers():
    vouchers = current_user.vouchers
    return render_template('my_vouchers.html', user=current_user, vouchers=vouchers)

def generate_voucher_code(length=10):
    letters_and_digits = string.ascii_letters + string.digits
    return ''.join(random.choice(letters_and_digits) for i in range(length))

stripe.api_key = 'sk_test_51P9F8mAHsD0ooQchiY1nEJu6Q5jpeRG1lvI8JqL0eHXuLbjDdgsZR9ai5TDU6h7qWcBk0EvKWTaWY02K4AIRoB9m00oxqHLQ4m'

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
@views.route('/admin-payments', methods=['GET', 'POST'])
@login_required
def admin_payments():
    if current_user.role != 'admin':
        flash('Brak dostępu', 'error')
        return redirect(url_for('views.home'))

    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename
            file_extension = os.path.splitext(filename)[1].lower()
            if file_extension != '.xlsx':
                flash("Akceptowane formaty plików to tylko XLSX", 'error')
                return redirect(url_for('views.admin_payments'))

            filename = secure_filename(filename)
            file_path = os.path.join('uploads', filename)
            file.save(file_path)

            try:
                wb = load_workbook(file_path)
                ws = wb.active
            except Exception as e:
                flash(f"Wystąpił błąd podczas wczytywania pliku: {e}", 'error')
                return redirect(url_for('views.admin_payments'))

            # Pobierz nagłówki
            actual_columns = [cell.value for cell in ws[1]]
            expected_columns = ['Email', 'Description', 'Amount', 'Due Date']

            # Sprawdzenie kolumn
            if not all(column in actual_columns for column in expected_columns):
                flash(f"Nazwy kolumn są niepoprawne, prosimy pobrać plik przykładowy. Znalezione kolumny: {', '.join(actual_columns)}. Wymagane kolumny: {', '.join(expected_columns)}.", 'error')
                return redirect(url_for('views.admin_payments'))

            # Przetwarzanie danych
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(actual_columns, row))
                user = User.query.filter_by(email=data['Email']).first()
                if user:
                    try:
                        due_date_str = data['Due Date'].strftime('%Y-%m-%d') if isinstance(data['Due Date'], datetime) else data['Due Date']
                        payment = Payment(
                            user_uid=user.uid,
                            description=data['Description'],
                            amount=data['Amount'],
                            due_date=datetime.strptime(due_date_str, '%Y-%m-%d')
                        )
                        db.session.add(payment)
                    except Exception as e:
                        flash(f"Błąd podczas przetwarzania wiersza: {e}", 'error')
                        return redirect(url_for('views.admin_payments'))
            db.session.commit()
            flash('Płatności zostały dodane', 'success')
            return redirect(url_for('views.admin_payments'))

    return render_template('admin_payments.html', user=current_user)

@views.route('/upload-payments', methods=['GET', 'POST'])
@login_required
def upload_payments():
    if current_user.role != 'admin':
        flash('Brak dostępu', 'error')
        return redirect(url_for('views.home'))

    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename
            file_extension = os.path.splitext(filename)[1].lower()
            if file_extension != '.xlsx':
                flash("Akceptowane formaty plików to tylko XLSX", 'error')
                return redirect(url_for('views.upload_payments'))

            filename = secure_filename(filename)
            file_path = os.path.join('uploads', filename)
            file.save(file_path)

            try:
                wb = load_workbook(file_path)
                ws = wb.active
            except Exception as e:
                flash(f"Wystąpił błąd podczas wczytywania pliku: {e}", 'error')
                return redirect(url_for('views.upload_payments'))

            # Pobierz nagłówki
            actual_columns = [cell.value for cell in ws[1]]
            expected_columns = ['Email', 'Description', 'Amount', 'Due Date']

            # Sprawdzenie kolumn
            if not all(column in actual_columns for column in expected_columns):
                flash(f"Nazwy kolumn są niepoprawne, prosimy pobrać plik przykładowy. Znalezione kolumny: {', '.join(actual_columns)}. Wymagane kolumny: {', '.join(expected_columns)}.", 'error')
                return redirect(url_for('views.upload_payments'))

            # Przetwarzanie danych
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(actual_columns, row))
                user = User.query.filter_by(email=data['Email']).first()
                if user:
                    try:
                        due_date_str = data['Due Date'].strftime('%Y-%m-%d') if isinstance(data['Due Date'], datetime) else data['Due Date']
                        payment = Payment(
                            user_uid=user.uid,
                            description=data['Description'],
                            amount=data['Amount'],
                            due_date=datetime.strptime(due_date_str, '%Y-%m-%d')
                        )
                        db.session.add(payment)
                    except Exception as e:
                        flash(f"Błąd podczas przetwarzania wiersza: {e}", 'error')
                        return redirect(url_for('views.upload_payments'))
            db.session.commit()
            flash('Płatności zostały dodane', 'success')
            return redirect(url_for('views.home'))

    return render_template('upload_payments.html', user=current_user)

@views.route('/payments')
@login_required
def payments():
    user_payments = Payment.query.filter_by(user_uid=current_user.uid).all()
    return render_template('user_payments.html', user=current_user, payments=user_payments)


@views.route('/create-checkout-session', methods=['POST'])
@login_required
def create_checkout_session():
    payment_id = request.form['payment_id']
    payment = Payment.query.get(payment_id)
    if payment and not payment.paid:
        session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'pln',
                    'product_data': {
                        'name': payment.description,
                    },
                    'unit_amount': int(payment.amount * 100),
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url=url_for('views.success', _external=True) + '?session_id={CHECKOUT_SESSION_ID}',
            cancel_url=url_for('views.cancel', _external=True),
        )
        return redirect(session.url, code=303)
    flash('Błąd podczas tworzenia sesji płatności')
    return redirect(url_for('views.payments'))

@views.route('/success')
@login_required
def success():
    return render_template('success.html', user=current_user)

@views.route('/cancel')
@login_required
def cancel():
    return render_template('cancel.html', user=current_user)


@views.route('/download-sample')
@login_required
def download_sample():
    if current_user.role != 'admin':
        flash('Brak dostępu')
        return redirect(url_for('views.home'))
    return send_from_directory(directory='.', path='sample.xlsx', as_attachment=True)